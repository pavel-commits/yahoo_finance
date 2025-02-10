import openpyxl
import requests
import re
import logging
import time
import json
import os.path
from bs4 import BeautifulSoup
# from datetime import datetime
from yfinance import Ticker
import warnings
from pandas import Timestamp

headers = {
    "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.3 Safari/605.1.15"
}

logging.basicConfig(filename="parser.log", level=logging.DEBUG, format="%(asctime)s:%(levelname)s:%(message)s")
warnings.filterwarnings("ignore")


class Parser:
    def __init__(self, AAPL, config, cookie, crumb):
        self.AAPL = AAPL
        self.timeout = config.get("timeout")
        self.cookie = cookie

        self.url_Income_Statement = f"https://finance.yahoo.com/quote/{AAPL}/financials?p={AAPL}"
        self.url_Balance_Sheet = f"https://finance.yahoo.com/quote/{AAPL}/balance-sheet?p={AAPL}"
        self.url_Cash_Flow = f"https://finance.yahoo.com/quote/{AAPL}/cash-flow?p={AAPL}"
        self.url_Corporate = f"https://finance.yahoo.com/quote/{AAPL}/profile?p={AAPL}"
        self.url_Api = f"https://query2.finance.yahoo.com/v10/finance/quoteSummary/{AAPL}?crumb={crumb}&modules=assetProfile%2CesgScores%2CfinancialData%2CdefaultKeyStatistics%2CsummaryDetail%2CearningsHistory%2CearningsTrend"

        # self.url_Api = f"https://query2.finance.yahoo.com/v10/finance/quoteSummary/{AAPL}?modules=assetProfile%2CsummaryProfile%2CsummaryDetail%2CesgScores%2Cprice%2CincomeStatementHistory%2CincomeStatementHistoryQuarterly%2CbalanceSheetHistory%2CbalanceSheetHistoryQuarterly%2CcashflowStatementHistory%2CcashflowStatementHistoryQuarterly%2CdefaultKeyStatistics%2CfinancialData%2CcalendarEvents%2CsecFilings%2CrecommendationTrend%2CupgradeDowngradeHistory%2CinstitutionOwnership%2CfundOwnership%2CmajorDirectHolders%2CmajorHoldersBreakdown%2CinsiderTransactions%2CinsiderHolders%2CnetSharePurchaseActivity%2Cearnings%2CearningsHistory%2CearningsTrend%2CindustryTrend%2CindexTrend%2CsectorTrend"
        # self.url_Api = f"https://query2.finance.yahoo.com/v10/finance/quoteSummary/{AAPL}?modules=assetProfile%2CesgScores%2CfinancialData%2CdefaultKeyStatistics%2CsummaryDetail%2CearningsHistory%2CearningsTrend"

        self.config = config
        self.data = {
            "Summary": {
                "1yTargetEst": None
            },
            "Profile": {
                "country": None,
                "website": None,
                "sector": None,
                "industry": None,
                "fullTimeEmployees": None,
                "longBusinessSummary": None,
                "corporateGovernance": None
            },
            "Statistics": {
                "averageDailyVolume3Month": None,
                "shares_short": None,
                "shortPercentOfFloat": None,
                "sharesShortPriorMonth": None,
                "dividendRate": None,
                "trailingAnnualDividendRate": None
            },
            "Analysis": {
                "numberOfAnalysts": {"CurrentQtr": None, "NextQtr": None, "CurrentYear": None, "NextYear": None},
                "earnings_avg": {"CurrentQtr": None, "NextQtr": None, "CurrentYear": None, "NextYear": None},
                "low": {"CurrentQtr": None, "NextQtr": None, "CurrentYear": None, "NextYear": None},
                "high": {"CurrentQtr": None, "NextQtr": None, "CurrentYear": None, "NextYear": None},
                "revenue_avg": {"CurrentQtr": None, "NextQtr": None, "CurrentYear": None, "NextYear": None},
                "salesGrowth": {"CurrentQtr": None, "NextQtr": None, "CurrentYear": None, "NextYear": None},
                "epsEstimate": None,
                "epsActual": None,
                "currentEstimate": None,
                "7daysAgo": None,
                "30daysAgo": None,
                "currentYear": None,
                "nextYear": None,
                "next5Years": None,
                "past5Years": None
            },
            "Sustainability": {
                "totalEsg": None,
                "environmentScore": None,
                "socialScore": None,
                "governanceScore": None,
                "highestControversy": None
            },
        }
        self.financials = {
            "Income_Statement": {
                "dates": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "total_revenue": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "operating_revenue": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "cost_of_revenue": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "gross_profit": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "operating_expense": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "selling_general_and_administrative": {"Annual": [None, None, None, None],
                                                       "Quarterly": [None, None, None, None]},
                "research_development": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "operating_income": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "net_non_operating_interest_income_expense": {"Annual": [None, None, None, None],
                                                              "Quarterly": [None, None, None, None]},
                "interest_income_non_operating": {"Annual": [None, None, None, None],
                                                  "Quarterly": [None, None, None, None]},
                "interest_expense_non_operating": {"Annual": [None, None, None, None],
                                                   "Quarterly": [None, None, None, None]},
                "pretax_income": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "tax_provision": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "net_income": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "basic_eps": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "diluted_eps": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "basic_average_shares": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "diluted_average_shares": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "ebit": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "ebitda": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "normalized_ebitda": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "net_interest_income": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "non_interest_income": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "non_interest_expense": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "occupancy_and_equipment": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},

                "general_administrative_expense": {"Annual": [None, None, None, None],
                                                   "Quarterly": [None, None, None, None]},
                "selling_marketing_expense": {"Annual": [None, None, None, None],
                                              "Quarterly": [None, None, None, None]},
                "other_non_interest_expense": {"Annual": [None, None, None, None],
                                               "Quarterly": [None, None, None, None]},
                "normalized_income": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]}
            },
            "Balance_Sheet": {
                "dates": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "current_debt": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "long_term_debt": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "cash_cash_equivalents_short_term_investments": {"Annual": [None, None, None, None],
                                                                 "Quarterly": [None, None, None, None]},
                "common_stock_equity": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "net_debt": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "ordinary_shares_number": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "preferred_shares_number": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "treasury_shares_number": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]}
            },
            "Cash_Flow": {
                "dates": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "depreciation_amortization": {"Annual": [None, None, None, None],
                                              "Quarterly": [None, None, None, None]},
                "common_stock_issuance": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "common_stock_payments": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "preferred_stock_issuance": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "preferred_stock_payments": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "cash_dividends_paid": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "investing_cash_flow": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]},
                "capital_expenditure": {"Annual": [None, None, None, None], "Quarterly": [None, None, None, None]}
            },
            "currency": None
        }

        # print(f"{AAPL}: START TIME", datetime.now().strftime("%H:%M:%S"))
        logging.info(f"{AAPL}: START TIME")
        print(self.AAPL, end=" ")

        self.get_Api()
        # if config.get("white"):
        #     self.get_corporate_governanse()

        if self.config.get("yellow"):
            self.get_Income_Statement()
            time.sleep(self.timeout)

            self.get_Balance_Sheet()
            time.sleep(self.timeout)

            self.get_Cash_Flow()

        # print(f"{AAPL}: END TIME", datetime.now())
        logging.info(f"{AAPL}: END TIME")
        # print("--------------------------------------")
        print()

    def get_Api(self):
        try:
            print("1", end=" ")
            # response = requests.get(self.url_Api, headers=headers)
            response = requests.get(
                self.url_Api,
                headers=headers,
                cookies={self.cookie.name: self.cookie.value},
                allow_redirects=True
            )
            if response:
                response = response.json()
                if response.get("quoteSummary", {}).get("error") is not None:
                    print()
                    print(f"api error {self.url_Api}")
                    logging.debug(f"api error {self.url_Api}")
                    print(response.get("quoteSummary", {}).get("error", {}).get("code"), "==>",
                          response.get("quoteSummary", {}).get("description", {}).get("code"))
                    logging.debug(response.get("quoteSummary", {}).get("error", {}).get("code"), "==>",
                                  response.get("quoteSummary", {}).get("description", {}).get("code"))
                    return
            else:
                print()
                print(f"api error {self.url_Api}")
                logging.debug(f"api error {self.url_Api}")
                return
            # with open(f"api/{self.AAPL}.json", "w") as f:
            #     json.dump(response, f)
            json_data = response.get("quoteSummary").get("result")[0]
            self.financials["currency"] = json_data.get("financialData").get("financialCurrency")

            # "Profile" & "Sustainability"
            if self.config.get("white"):
                pr = json_data.get("assetProfile", {})

                self.data["Profile"]["country"] = pr.get("country", None)
                self.data["Profile"]["website"] = pr.get("website", None)
                self.data["Profile"]["sector"] = pr.get("sector", None)
                self.data["Profile"]["industry"] = pr.get("industry", None)
                self.data["Profile"]["fullTimeEmployees"] = pr.get("fullTimeEmployees", None)

                sus = json_data.get("esgScores", {})

                self.data["Sustainability"]["totalEsg"] = sus.get("totalEsg", {}).get("raw", None)
                self.data["Sustainability"]["environmentScore"] = sus.get("environmentScore", {}).get("raw", None)
                self.data["Sustainability"]["socialScore"] = sus.get("socialScore", {}).get("raw", None)
                self.data["Sustainability"]["governanceScore"] = sus.get("governanceScore", {}).get("raw", None)
                self.data["Sustainability"]["highestControversy"] = sus.get("highestControversy", None)

            # "Summary" & "Statistics"
            if self.config.get("yellow"):
                self.data["Summary"]["1yTargetEst"] = json_data.get("financialData", {}).get("targetMeanPrice", {}).get(
                    "raw", None)

                sd = json_data.get("summaryDetail", {})

                self.data["Statistics"]["averageDailyVolume3Month"] = sd.get("averageVolume", {}).get("raw", None)

                self.data["Statistics"]["dividendRate"] = sd.get("dividendRate", {}).get("raw", None)
                self.data["Statistics"]["trailingAnnualDividendRate"] = sd.get("trailingAnnualDividendRate", {}).get(
                    "raw",
                    None)

            if self.config.get("shares"):
                dks = json_data.get("defaultKeyStatistics", {})
                self.data["Statistics"]["sharesShortPriorMonth"] = dks.get("sharesShortPriorMonth", {}).get("raw", None)
                self.data["Statistics"]["shares_short"] = dks.get("sharesShort", {}).get("raw", None)
                self.data["Statistics"]["shortPercentOfFloat"] = dks.get("shortPercentOfFloat", {}).get("raw", None)

            if self.config.get("description"):
                pr = json_data.get("assetProfile", {})
                self.data["Profile"]["longBusinessSummary"] = pr.get("longBusinessSummary", None)

            # Analysys
            if self.config.get("green"):
                tr = json_data.get("earningsTrend", {}).get("trend", [{}, {}, {}, {}, {}, {}])
                trend_0 = tr[0]
                trend_1 = tr[1]
                trend_2 = tr[2]
                trend_3 = tr[3]

                # =--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
                ear_0 = trend_0.get("earningsEstimate", {})
                ear_1 = trend_1.get("earningsEstimate", {})
                ear_2 = trend_2.get("earningsEstimate", {})
                ear_3 = trend_3.get("earningsEstimate", {})

                # {"CurrentQtr": None, "NextQtr": None, "CurrentYear": None, "NextYear": None}
                self.data["Analysis"]["numberOfAnalysts"]["CurrentYear"] = ear_2.get("numberOfAnalysts", {}).get("raw",
                                                                                                                 None)
                self.data["Analysis"]["numberOfAnalysts"]["NextYear"] = ear_3.get("numberOfAnalysts", {}).get("raw",
                                                                                                              None)
                self.data["Analysis"]["numberOfAnalysts"]["CurrentQtr"] = ear_0.get("numberOfAnalysts", {}).get("raw",
                                                                                                                None)
                self.data["Analysis"]["numberOfAnalysts"]["NextQtr"] = ear_1.get("numberOfAnalysts", {}).get("raw",
                                                                                                             None)

                self.data["Analysis"]["earnings_avg"]["CurrentYear"] = ear_2.get("avg", {}).get("raw", None)
                self.data["Analysis"]["earnings_avg"]["NextYear"] = ear_3.get("avg", {}).get("raw", None)
                self.data["Analysis"]["earnings_avg"]["CurrentQtr"] = ear_0.get("avg", {}).get("raw", None)
                self.data["Analysis"]["earnings_avg"]["NextQtr"] = ear_1.get("avg", {}).get("raw", None)

                self.data["Analysis"]["low"]["CurrentYear"] = ear_2.get("low", {}).get("raw", None)
                self.data["Analysis"]["low"]["NextYear"] = ear_3.get("low", {}).get("raw", None)
                self.data["Analysis"]["low"]["CurrentQtr"] = ear_0.get("low", {}).get("raw", None)
                self.data["Analysis"]["low"]["NextQtr"] = ear_1.get("low", {}).get("raw", None)

                self.data["Analysis"]["high"]["CurrentYear"] = ear_2.get("high", {}).get("raw", None)
                self.data["Analysis"]["high"]["NextYear"] = ear_3.get("high", {}).get("raw", None)
                self.data["Analysis"]["high"]["CurrentQtr"] = ear_0.get("high", {}).get("raw", None)
                self.data["Analysis"]["high"]["NextQtr"] = ear_1.get("high", {}).get("raw", None)

                # =--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
                rev_0 = trend_0.get("revenueEstimate", {})
                rev_1 = trend_1.get("revenueEstimate", {})
                rev_2 = trend_2.get("revenueEstimate", {})
                rev_3 = trend_3.get("revenueEstimate", {})

                self.data["Analysis"]["revenue_avg"]["CurrentYear"] = rev_2.get("avg", {}).get("raw")
                self.data["Analysis"]["revenue_avg"]["NextYear"] = rev_3.get("avg", {}).get("raw")
                self.data["Analysis"]["revenue_avg"]["CurrentQtr"] = rev_0.get("avg", {}).get("raw")
                self.data["Analysis"]["revenue_avg"]["NextQtr"] = rev_1.get("avg", {}).get("raw")

                self.data["Analysis"]["salesGrowth"]["CurrentYear"] = rev_2.get("growth", {}).get("raw")
                self.data["Analysis"]["salesGrowth"]["NextYear"] = rev_3.get("growth", {}).get("raw")
                self.data["Analysis"]["salesGrowth"]["CurrentQtr"] = rev_0.get("growth", {}).get("raw")
                self.data["Analysis"]["salesGrowth"]["NextQtr"] = rev_1.get("growth", {}).get("raw")

                # =--=-=-=-=-=-=-=-=-=-=-=-=-=-=-
                hist = json_data.get("earningsHistory", {}).get("history", [{}])[0]
                self.data["Analysis"]["epsEstimate"] = hist.get("epsEstimate", {}).get("raw", None)
                self.data["Analysis"]["epsActual"] = hist.get("epsActual", {}).get("raw", None)

                # =--=-=-=-=-=-=-=-=-=-=-=-=-=-=-
                eps_0 = trend_0.get("epsTrend", {})

                self.data["Analysis"]["currentEstimate"] = eps_0.get("current", {}).get("raw", None)
                self.data["Analysis"]["7daysAgo"] = eps_0.get("7daysAgo", {}).get("raw", None)
                self.data["Analysis"]["30daysAgo"] = eps_0.get("30daysAgo", {}).get("raw", None)

                # =--=-=-=-=-=-=-=-=-=-=-=-=-=-=-

                self.data["Analysis"]["currentYear"] = tr[2].get("growth", {}).get("raw", None)
                self.data["Analysis"]["nextYear"] = tr[3].get("growth", {}).get("raw", None)
                self.data["Analysis"]["next5Years"] = tr[4].get("growth", {}).get("raw", None)
                self.data["Analysis"]["past5Years"] = tr[5].get("growth", {}).get("raw", None)
        except Exception as e:
            print(f"error get_api {self.url_Api}")
            print(e)
            logging.error(f"{e.__class__.__name__}, {e}, {self.url_Api}")

    def get_corporate_governanse(self):
        try:
            print("2", end=" ")
            logging.debug(f"Make response to corporate_governanse {self.url_Corporate}")
            response = requests.get(self.url_Corporate, headers=headers)

            soup = BeautifulSoup(response.text, "html.parser")
            container = soup.find("section", class_="corporate-governance-container").find("div",
                                                                                           class_="Mt(20px)").find(
                "span", text=re.compile(r'while')).text
            container = container[container.find("while a "):].split()
            self.data["Profile"]["corporateGovernance"] = int(container[2])
        except Exception as e:
            print("error corporate_governanse", self.url_Corporate)
            logging.error(f"{e.__class__.__name__}, {e}, {self.url_Corporate}")

    def get_Income_Statement(self, times=None):
        try:
            print("2", end=" ")
            ticker = Ticker(self.AAPL)
            ddd = ticker.get_income_stmt(as_dict=True)
            ddq = ticker.get_income_stmt(as_dict=True, freq="quarterly")
            keys = [g for g in ddd.keys()]
            keys_q = [g for g in ddq.keys()]
            # print(ddd)
            for i in range(min(len(keys), 4)):
                self.financials["Income_Statement"]["dates"]["Annual"][i] = keys[i]

                self.financials["Income_Statement"]["total_revenue"]["Annual"][i] = ddd[keys[i]].get("TotalRevenue")
                self.financials["Income_Statement"]["operating_revenue"]["Annual"][i] = ddd[keys[i]].get(
                    "OperatingRevenue")
                self.financials["Income_Statement"]["cost_of_revenue"]["Annual"][i] = ddd[keys[i]].get("CostOfRevenue")
                self.financials["Income_Statement"]["gross_profit"]["Annual"][i] = ddd[keys[i]].get("GrossProfit")
                self.financials["Income_Statement"]["operating_expense"]["Annual"][i] = ddd[keys[i]].get(
                    "OperatingExpense")
                self.financials["Income_Statement"]["selling_general_and_administrative"]["Annual"][i] = ddd[
                    keys[i]].get("SellingGeneralAndAdministration")
                self.financials["Income_Statement"]["research_development"]["Annual"][i] = ddd[keys[i]].get(
                    "ResearchAndDevelopment")
                self.financials["Income_Statement"]["operating_income"]["Annual"][i] = ddd[keys[i]].get(
                    "OperatingIncome")
                self.financials["Income_Statement"]["net_non_operating_interest_income_expense"]["Annual"][i] = ddd[
                    keys[i]].get("NetNonOperatingInterestIncomeExpense")
                self.financials["Income_Statement"]["interest_income_non_operating"]["Annual"][i] = ddd[keys[i]].get(
                    "InterestIncomeNonOperating")
                self.financials["Income_Statement"]["interest_expense_non_operating"]["Annual"][i] = ddd[keys[i]].get(
                    "InterestExpenseNonOperating")
                self.financials["Income_Statement"]["pretax_income"]["Annual"][i] = ddd[keys[i]].get("PretaxIncome")
                self.financials["Income_Statement"]["tax_provision"]["Annual"][i] = ddd[keys[i]].get("TaxProvision")
                self.financials["Income_Statement"]["net_income"]["Annual"][i] = ddd[keys[i]].get("NetIncome")
                self.financials["Income_Statement"]["basic_eps"]["Annual"][i] = ddd[keys[i]].get("BasicEPS")
                self.financials["Income_Statement"]["diluted_eps"]["Annual"][i] = ddd[keys[i]].get("DilutedEPS")
                self.financials["Income_Statement"]["basic_average_shares"]["Annual"][i] = ddd[keys[i]].get(
                    "BasicAverageShares")
                self.financials["Income_Statement"]["diluted_average_shares"]["Annual"][i] = ddd[keys[i]].get(
                    "DilutedAverageShares")
                self.financials["Income_Statement"]["ebit"]["Annual"][i] = ddd[keys[i]].get("EBIT")
                self.financials["Income_Statement"]["ebitda"]["Annual"][i] = ddd[keys[i]].get("EBITDA")
                self.financials["Income_Statement"]["normalized_ebitda"]["Annual"][i] = ddd[keys[i]].get(
                    "NormalizedEBITDA")
                self.financials["Income_Statement"]["net_interest_income"]["Annual"][i] = ddd[keys[i]].get(
                    "NetInterestIncome")
                self.financials["Income_Statement"]["general_administrative_expense"]["Annual"][i] = ddd[keys[i]].get(
                    "GeneralAndAdministrativeExpense")
                self.financials["Income_Statement"]["normalized_income"]["Annual"][i] = ddd[keys[i]].get(
                    "NormalizedIncome")
                self.financials["Income_Statement"]["selling_marketing_expense"]["Annual"][i] = ddd[keys[i]].get(
                    "SellingAndMarketingExpense")

                # -=-=-=-=-=-==-=-=--=-=-=
                self.financials["Income_Statement"]["non_interest_income"]["Annual"][i] = ddd[keys[i]].get(
                    "NonInterestIncome")
                self.financials["Income_Statement"]["non_interest_expense"]["Annual"][i] = ddd[keys[i]].get(
                    "NonInterestExpense")
                self.financials["Income_Statement"]["occupancy_and_equipment"]["Annual"][i] = ddd[keys[i]].get(
                    "OccupancyAndEquipment")
                self.financials["Income_Statement"]["other_non_interest_expense"]["Annual"][i] = ddd[keys[i]].get(
                    "OtherNonInterestExpense")

            for i in range(min(len(keys_q), 4)):
                self.financials["Income_Statement"]["dates"]["Quarterly"][i] = keys_q[i]
                self.financials["Income_Statement"]["total_revenue"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "TotalRevenue")
                self.financials["Income_Statement"]["operating_revenue"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "OperatingRevenue")
                self.financials["Income_Statement"]["cost_of_revenue"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "CostOfRevenue")
                self.financials["Income_Statement"]["gross_profit"]["Quarterly"][i] = ddq[keys_q[i]].get("GrossProfit")
                self.financials["Income_Statement"]["operating_expense"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "OperatingExpense")
                self.financials["Income_Statement"]["selling_general_and_administrative"]["Quarterly"][i] = ddq[
                    keys_q[i]].get("SellingGeneralAndAdministration")
                self.financials["Income_Statement"]["research_development"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "ResearchAndDevelopment")
                self.financials["Income_Statement"]["operating_income"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "OperatingIncome")
                self.financials["Income_Statement"]["net_non_operating_interest_income_expense"]["Quarterly"][i] = ddq[
                    keys_q[i]].get("NetNonOperatingInterestIncomeExpense")
                self.financials["Income_Statement"]["interest_income_non_operating"]["Quarterly"][i] = ddq[
                    keys_q[i]].get("InterestIncomeNonOperating")
                self.financials["Income_Statement"]["interest_expense_non_operating"]["Quarterly"][i] = ddq[
                    keys_q[i]].get("InterestExpenseNonOperating")
                self.financials["Income_Statement"]["pretax_income"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "PretaxIncome")
                self.financials["Income_Statement"]["tax_provision"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "TaxProvision")
                self.financials["Income_Statement"]["net_income"]["Quarterly"][i] = ddq[keys_q[i]].get("NetIncome")
                self.financials["Income_Statement"]["basic_eps"]["Quarterly"][i] = ddq[keys_q[i]].get("BasicEPS")
                self.financials["Income_Statement"]["diluted_eps"]["Quarterly"][i] = ddq[keys_q[i]].get("DilutedEPS")
                self.financials["Income_Statement"]["basic_average_shares"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "BasicAverageShares")
                self.financials["Income_Statement"]["diluted_average_shares"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "DilutedAverageShares")
                self.financials["Income_Statement"]["ebit"]["Quarterly"][i] = ddq[keys_q[i]].get("EBIT")
                self.financials["Income_Statement"]["ebitda"]["Quarterly"][i] = ddq[keys_q[i]].get("EBITDA")
                self.financials["Income_Statement"]["normalized_ebitda"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "NormalizedEBITDA")
                self.financials["Income_Statement"]["net_interest_income"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "NetInterestIncome")
                self.financials["Income_Statement"]["general_administrative_expense"]["Quarterly"][i] = ddq[
                    keys_q[i]].get("GeneralAndAdministrativeExpense")
                self.financials["Income_Statement"]["normalized_income"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "NormalizedIncome")
                self.financials["Income_Statement"]["selling_marketing_expense"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "SellingAndMarketingExpense")

                # -=-=-=-=-=-==-=-=--=-=-=
                self.financials["Income_Statement"]["non_interest_income"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "NonInterestIncome")
                self.financials["Income_Statement"]["non_interest_expense"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "NonInterestExpense")
                self.financials["Income_Statement"]["occupancy_and_equipment"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "OccupancyAndEquipment")
                self.financials["Income_Statement"]["other_non_interest_expense"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "OtherNonInterestExpense")

        except Exception as e:
            print(f"error Income_Statement {self.url_Income_Statement}")
            print(e)
            logging.error(f"{e.__class__.__name__}, {e}, {self.url_Income_Statement}")
            if times:
                return
            time.sleep(self.timeout)
            print("restarting get Income Statement...")
            logging.debug("restarting get Income Statement...")
            try:
                self.get_Income_Statement(times=1)
            except Exception as e:
                print("error Income_Statement", self.url_Income_Statement)
                logging.error(f"{e.__class__.__name__}, {e}, {self.url_Income_Statement}")

    def get_Balance_Sheet(self, times=None):
        try:
            print("3", end=" ")
            ticker = Ticker(self.AAPL)
            ddd = ticker.get_balance_sheet(as_dict=True)
            ddq = ticker.get_balance_sheet(as_dict=True, freq="quarterly")
            keys = [g for g in ddd.keys()]
            keys_q = [g for g in ddq.keys()]

            for i in range(min(len(keys), 4)):
                self.financials["Balance_Sheet"]["dates"]["Annual"][i] = keys[i]
                self.financials["Balance_Sheet"]["current_debt"]["Annual"][i] = ddd[keys[i]].get("CurrentDebt")
                self.financials["Balance_Sheet"]["cash_cash_equivalents_short_term_investments"]["Annual"][i] = ddd[
                    keys[i]].get("CashCashEquivalentsAndShortTermInvestments")
                self.financials["Balance_Sheet"]["net_debt"]["Annual"][i] = ddd[keys[i]].get("NetDebt")
                self.financials["Balance_Sheet"]["ordinary_shares_number"]["Annual"][i] = ddd[keys[i]].get(
                    "OrdinarySharesNumber")
                self.financials["Balance_Sheet"]["long_term_debt"]["Annual"][i] = ddd[keys[i]].get("LongTermDebt")
                self.financials["Balance_Sheet"]["common_stock_equity"]["Annual"][i] = ddd[keys[i]].get(
                    "CommonStockEquity")

                self.financials["Balance_Sheet"]["preferred_shares_number"]["Annual"][i] = ddd[keys[i]].get(
                    "PreferredSharesNumber")
                self.financials["Balance_Sheet"]["treasury_shares_number"]["Annual"][i] = ddd[keys[i]].get(
                    "TreasurySharesNumber")
            for i in range(min(len(keys_q), 4)):
                self.financials["Balance_Sheet"]["dates"]["Quarterly"][i] = keys_q[i]
                self.financials["Balance_Sheet"]["current_debt"]["Quarterly"][i] = ddq[keys_q[i]].get("CurrentDebt")
                self.financials["Balance_Sheet"]["cash_cash_equivalents_short_term_investments"]["Quarterly"][i] = ddq[
                    keys_q[i]].get("CashCashEquivalentsAndShortTermInvestments")
                self.financials["Balance_Sheet"]["net_debt"]["Quarterly"][i] = ddq[keys_q[i]].get("NetDebt")
                self.financials["Balance_Sheet"]["ordinary_shares_number"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "OrdinarySharesNumber")
                self.financials["Balance_Sheet"]["long_term_debt"]["Quarterly"][i] = ddq[keys_q[i]].get("LongTermDebt")
                self.financials["Balance_Sheet"]["common_stock_equity"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "CommonStockEquity")
                self.financials["Balance_Sheet"]["preferred_shares_number"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "PreferredSharesNumber")
                self.financials["Balance_Sheet"]["treasury_shares_number"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "TreasurySharesNumber")
        except Exception as e:
            print("error Balance_Sheet", self.url_Balance_Sheet)
            print(e)
            logging.error(e.__class__.__name__, e, self.url_Balance_Sheet)
            if times:
                return
            time.sleep(self.timeout)
            print("restarting get Balance Sheet...")
            logging.debug("restarting get Balance Sheet...")
            try:
                self.get_Balance_Sheet(times=1)
            except Exception as e:
                print("error Balance_Sheet", self.url_Balance_Sheet)
                logging.error(e.__class__.__name__, e, self.url_Balance_Sheet)

    def get_Cash_Flow(self, times=None):
        try:
            print("4", end=" ")
            ticker = Ticker(self.AAPL)
            ddd = ticker.get_cash_flow(as_dict=True)
            ddq = ticker.get_cash_flow(as_dict=True, freq="quarterly")
            keys = [g for g in ddd.keys()]
            keys_q = [g for g in ddq.keys()]

            for i in range(min(len(keys), 4)):
                self.financials["Cash_Flow"]["dates"]["Annual"][i] = keys[i]
                self.financials["Cash_Flow"]["depreciation_amortization"]["Annual"][i] = ddd[keys[i]].get(
                    "DepreciationAndAmortization")
                self.financials["Cash_Flow"]["common_stock_issuance"]["Annual"][i] = ddd[keys[i]].get(
                    "CommonStockIssuance")
                self.financials["Cash_Flow"]["common_stock_payments"]["Annual"][i] = ddd[keys[i]].get(
                    "CommonStockPayments")
                self.financials["Cash_Flow"]["cash_dividends_paid"]["Annual"][i] = ddd[keys[i]].get("CashDividendsPaid")
                self.financials["Cash_Flow"]["investing_cash_flow"]["Annual"][i] = ddd[keys[i]].get("InvestingCashFlow")

                self.financials["Cash_Flow"]["preferred_stock_issuance"]["Annual"][i] = ddd[keys[i]].get(
                    "PreferredStockIssuance")
                self.financials["Cash_Flow"]["preferred_stock_payments"]["Annual"][i] = ddd[keys[i]].get(
                    "PreferredStockPayments")
                self.financials["Cash_Flow"]["capital_expenditure"]["Annual"][i] = ddd[keys[i]].get(
                    "CapitalExpenditure")

            for i in range(min(len(keys_q), 4)):
                self.financials["Cash_Flow"]["dates"]["Quarterly"][i] = keys_q[i]
                self.financials["Cash_Flow"]["depreciation_amortization"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "DepreciationAndAmortization")
                self.financials["Cash_Flow"]["common_stock_issuance"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "CommonStockIssuance")
                self.financials["Cash_Flow"]["common_stock_payments"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "CommonStockPayments")
                self.financials["Cash_Flow"]["cash_dividends_paid"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "CashDividendsPaid")
                self.financials["Cash_Flow"]["investing_cash_flow"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "InvestingCashFlow")

                self.financials["Cash_Flow"]["preferred_stock_issuance"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "PreferredStockIssuance")
                self.financials["Cash_Flow"]["preferred_stock_payments"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "PreferredStockPayments")
                self.financials["Cash_Flow"]["capital_expenditure"]["Quarterly"][i] = ddq[keys_q[i]].get(
                    "CapitalExpenditure")

        except Exception as e:
            print("error Cash_Flow", self.url_Cash_Flow)
            print(e)
            logging.error(f"{e.__class__.__name__}, {e}, {self.url_Cash_Flow}")
            if times:
                return
            time.sleep(self.timeout)
            print("restarting get Cash Flow...")
            logging.debug("restarting get Cash Flow...")
            try:
                self.get_Cash_Flow(times=1)
            except Exception as e:
                print("error Cash_Flow", self.url_Cash_Flow)
                logging.error(f"{e.__class__.__name__}, {e}, {self.url_Cash_Flow}")


def xlsx_main(file):
    book = openpyxl.Workbook()
    sheet = book.active
    sheet["A1"] = "Ticker"
    sheet["B1"] = "Summary"
    sheet["C1"] = "Profile"
    sheet["J1"] = "Statistics"
    sheet["P1"] = "Analysis"
    sheet["AW1"] = "Sustainability"

    sheet["B2"] = "1yTargetEst"

    sheet["C2"] = "country"
    sheet["D2"] = "website"
    sheet["E2"] = "sector"
    sheet["F2"] = "industry"
    sheet["G2"] = "fullTimeEmployees"
    sheet["H2"] = "longBusinessSummary"
    sheet["I2"] = "corporateGovernance"

    sheet["J2"] = "averageDailyVolume3Month"
    sheet["K2"] = "shares_short"
    sheet["L2"] = "shortPercentOfFloat"
    sheet["M2"] = "sharesShortPriorMonth"
    sheet["N2"] = "dividendRate"
    sheet["O2"] = "trailingAnnualDividendRate"

    sheet["P2"] = "numberOfAnalysts"
    sheet["T2"] = "earnings_avg"
    sheet["X2"] = "low"
    sheet["AB2"] = "high"
    sheet["AF2"] = "revenue_avg"
    sheet["AJ2"] = "salesGrowth"

    sheet["P3"] = "Qtr_cur"
    sheet["Q3"] = "Qtr_next"
    sheet["R3"] = "Year_cur"
    sheet["S3"] = "Year_next"
    sheet["T3"] = "Qtr_cur"
    sheet["U3"] = "Qtr_next"
    sheet["V3"] = "Year_cur"
    sheet["W3"] = "Year_next"
    sheet["X3"] = "Qtr_cur"
    sheet["Y3"] = "Qtr_next"
    sheet["Z3"] = "Year_cur"
    sheet["AA3"] = "Year_next"
    sheet["AB3"] = "Qtr_cur"
    sheet["AC3"] = "Qtr_next"
    sheet["AD3"] = "Year_cur"
    sheet["AE3"] = "Year_next"
    sheet["AF3"] = "Qtr_cur"
    sheet["AG3"] = "Qtr_next"
    sheet["AH3"] = "Year_cur"
    sheet["AI3"] = "Year_next"
    sheet["AJ3"] = "Qtr_cur"
    sheet["AK3"] = "Qtr_next"
    sheet["AL3"] = "Year_cur"
    sheet["AM3"] = "Year_next"

    sheet["AN2"] = "epsEstimate"
    sheet["AO2"] = "epsActual"
    sheet["AP2"] = "currentEstimate"
    sheet["AQ2"] = "7daysAgo"
    sheet["AR2"] = "30daysAgo"
    sheet["AS2"] = "currentYear"
    sheet["AT2"] = "nextYear"
    sheet["AU2"] = "next5Years"
    sheet["AV2"] = "past5Years"

    sheet["AW2"] = "totalEsg"
    sheet["AX2"] = "environmentScore"
    sheet["AY2"] = "socialScore"
    sheet["AZ2"] = "governanceScore"
    sheet["BA2"] = "highestControversy"

    book.save(file)
    book.close()
    print(f"File '{file}' is created")
    logging.info(f"File '{file}' is created")


def xlsx_main_fin(file):
    book = openpyxl.Workbook()
    sheet = book.active
    sheet["A1"] = "Ticker"
    sheet["B1"] = "Currency"
    sheet["C1"] = "Income_Statement"
    sheet["II1"] = "Balance_Sheet"
    sheet["LC1"] = "Cash_Flow"

    keys = ['dates', 'total_revenue', 'operating_revenue', 'cost_of_revenue', 'gross_profit', 'operating_expense',
            'selling_general_and_administrative', 'research_development', 'operating_income',
            'net_non_operating_interest_income_expense', 'interest_income_non_operating',
            'interest_expense_non_operating', 'pretax_income', 'tax_provision', 'net_income', 'basic_eps',
            'diluted_eps', 'basic_average_shares', 'diluted_average_shares', 'ebit', 'ebitda', 'normalized_ebitda',
            'net_interest_income', 'non_interest_income', 'non_interest_expense', 'occupancy_and_equipment',
            'general_administrative_expense', 'selling_marketing_expense', 'other_non_interest_expense',
            'normalized_income', 'dates', 'current_debt', 'long_term_debt',
            'cash_cash_equivalents_short_term_investments',
            'common_stock_equity', 'net_debt', 'ordinary_shares_number', 'preferred_shares_number',
            'treasury_shares_number', 'dates', 'depreciation_amortization', 'common_stock_issuance',
            'common_stock_payments',
            'preferred_stock_issuance', 'preferred_stock_payments', 'cash_dividends_paid', 'investing_cash_flow',
            'capital_expenditure']

    for i in range(len(keys)):
        sheet.cell(row=2, column=i * 8 + 3).value = keys[i]

        sheet.cell(row=3, column=i * 8 + 3).value = "date1"
        sheet.cell(row=3, column=i * 8 + 4).value = "date2"
        sheet.cell(row=3, column=i * 8 + 5).value = "date3"
        sheet.cell(row=3, column=i * 8 + 6).value = "date4"

        sheet.cell(row=3, column=i * 8 + 7).value = "Q1"
        sheet.cell(row=3, column=i * 8 + 8).value = "Q2"
        sheet.cell(row=3, column=i * 8 + 9).value = "Q3"
        sheet.cell(row=3, column=i * 8 + 10).value = "Q4"

    book.save(file)
    book.close()
    print(f"File '{file}' is created")
    print()
    logging.info(f"File '{file}' is created")


def to_xlsx(file, stock, data, row):
    try:
        book = openpyxl.open(file)
        sheet = book.active
        sheet[f"A{row}"] = stock

        sheet[f"B{row}"] = data.get("Summary").get("1yTargetEst")

        sheet[f"C{row}"] = data.get("Profile").get("country")
        sheet[f"D{row}"] = data.get("Profile").get("website")
        sheet[f"E{row}"] = data.get("Profile").get("sector")
        sheet[f"F{row}"] = data.get("Profile").get("industry")
        sheet[f"G{row}"] = data.get("Profile").get("fullTimeEmployees")
        sheet[f"H{row}"] = data.get("Profile").get("longBusinessSummary")
        sheet[f"I{row}"] = data.get("Profile").get("corporateGovernance")

        sheet[f"J{row}"] = data.get("Statistics").get("averageDailyVolume3Month")
        sheet[f"K{row}"] = data.get("Statistics").get("shares_short")
        sheet[f"L{row}"] = data.get("Statistics").get("shortPercentOfFloat")
        sheet[f"M{row}"] = data.get("Statistics").get("sharesShortPriorMonth")
        sheet[f"N{row}"] = data.get("Statistics").get("dividendRate")
        sheet[f"O{row}"] = data.get("Statistics").get("trailingAnnualDividendRate")

        sheet[f"P{row}"] = data.get("Analysis").get("numberOfAnalysts").get("CurrentQtr")
        sheet[f"Q{row}"] = data.get("Analysis").get("numberOfAnalysts").get("NextQtr")
        sheet[f"R{row}"] = data.get("Analysis").get("numberOfAnalysts").get("CurrentYear")
        sheet[f"S{row}"] = data.get("Analysis").get("numberOfAnalysts").get("NextYear")
        sheet[f"T{row}"] = data.get("Analysis").get("earnings_avg").get("CurrentQtr")
        sheet[f"U{row}"] = data.get("Analysis").get("earnings_avg").get("NextQtr")
        sheet[f"V{row}"] = data.get("Analysis").get("earnings_avg").get("CurrentYear")
        sheet[f"W{row}"] = data.get("Analysis").get("earnings_avg").get("NextYear")
        sheet[f"X{row}"] = data.get("Analysis").get("low").get("CurrentQtr")
        sheet[f"Y{row}"] = data.get("Analysis").get("low").get("NextQtr")
        sheet[f"Z{row}"] = data.get("Analysis").get("low").get("CurrentYear")
        sheet[f"AA{row}"] = data.get("Analysis").get("low").get("NextYear")
        sheet[f"AB{row}"] = data.get("Analysis").get("high").get("CurrentQtr")
        sheet[f"AC{row}"] = data.get("Analysis").get("high").get("NextQtr")
        sheet[f"AD{row}"] = data.get("Analysis").get("high").get("CurrentYear")
        sheet[f"AE{row}"] = data.get("Analysis").get("high").get("NextYear")
        sheet[f"AF{row}"] = data.get("Analysis").get("revenue_avg").get("CurrentQtr")
        sheet[f"AG{row}"] = data.get("Analysis").get("revenue_avg").get("NextQtr")
        sheet[f"AH{row}"] = data.get("Analysis").get("revenue_avg").get("CurrentYear")
        sheet[f"AI{row}"] = data.get("Analysis").get("revenue_avg").get("NextYear")
        sheet[f"AJ{row}"] = data.get("Analysis").get("salesGrowth").get("CurrentQtr")
        sheet[f"AK{row}"] = data.get("Analysis").get("salesGrowth").get("NextQtr")
        sheet[f"AL{row}"] = data.get("Analysis").get("salesGrowth").get("CurrentYear")
        sheet[f"AM{row}"] = data.get("Analysis").get("salesGrowth").get("NextYear")

        sheet[f"AN{row}"] = data.get("Analysis").get("epsEstimate")
        sheet[f"AO{row}"] = data.get("Analysis").get("epsActual")
        sheet[f"AP{row}"] = data.get("Analysis").get("currentEstimate")
        sheet[f"AQ{row}"] = data.get("Analysis").get("7daysAgo")
        sheet[f"AR{row}"] = data.get("Analysis").get("30daysAgo")
        sheet[f"AS{row}"] = data.get("Analysis").get("currentYear")
        sheet[f"AT{row}"] = data.get("Analysis").get("nextYear")
        sheet[f"AU{row}"] = data.get("Analysis").get("next5Years")
        sheet[f"AV{row}"] = data.get("Analysis").get("past5Years")

        sheet[f"AW{row}"] = data.get("Sustainability").get("totalEsg")
        sheet[f"AX{row}"] = data.get("Sustainability").get("environmentScore")
        sheet[f"AY{row}"] = data.get("Sustainability").get("socialScore")
        sheet[f"AZ{row}"] = data.get("Sustainability").get("governanceScore")
        sheet[f"BA{row}"] = data.get("Sustainability").get("highestControversy")

        book.save(file)
        book.close()
    except Exception as e:
        print(row)
        print(e)
        logging.error(f"to_xlsx() error, row: {row}. {e.__class__.__name__}, {e}")


def to_xlsx_fin(file, stock, data, row):
    try:
        book = openpyxl.open(file)
        sheet = book.active
        sheet[f"A{row}"] = stock

        income = ['dates', 'total_revenue', 'operating_revenue', 'cost_of_revenue', 'gross_profit', 'operating_expense',
                  'selling_general_and_administrative', 'research_development', 'operating_income',
                  'net_non_operating_interest_income_expense', 'interest_income_non_operating',
                  'interest_expense_non_operating', 'pretax_income', 'tax_provision', 'net_income', 'basic_eps',
                  'diluted_eps', 'basic_average_shares', 'diluted_average_shares', 'ebit', 'ebitda',
                  'normalized_ebitda',
                  'net_interest_income', 'non_interest_income', 'non_interest_expense', 'occupancy_and_equipment',
                  'general_administrative_expense', 'selling_marketing_expense', 'other_non_interest_expense',
                  'normalized_income']
        bal = ['dates', 'current_debt', 'long_term_debt', 'cash_cash_equivalents_short_term_investments',
               'common_stock_equity', 'net_debt', 'ordinary_shares_number', 'preferred_shares_number',
               'treasury_shares_number']
        cash = ['dates', 'depreciation_amortization', 'common_stock_issuance', 'common_stock_payments',
                'preferred_stock_issuance', 'preferred_stock_payments', 'cash_dividends_paid', 'investing_cash_flow',
                'capital_expenditure']
        # ccc = f'[${data.get("currency")} ]#,0'
        sheet[f"B{row}"] = data.get("currency")
        for i in range(len(income)):
            sheet.cell(row=row, column=i * 8 + 3).value = data.get("Income_Statement").get(income[i]).get("Annual")[0]
            sheet.cell(row=row, column=i * 8 + 4).value = data.get("Income_Statement").get(income[i]).get("Annual")[1]
            sheet.cell(row=row, column=i * 8 + 5).value = data.get("Income_Statement").get(income[i]).get("Annual")[2]
            sheet.cell(row=row, column=i * 8 + 6).value = data.get("Income_Statement").get(income[i]).get("Annual")[3]

            sheet.cell(row=row, column=i * 8 + 7).value = data.get("Income_Statement").get(income[i]).get("Quarterly")[
                0]
            sheet.cell(row=row, column=i * 8 + 8).value = data.get("Income_Statement").get(income[i]).get("Quarterly")[
                1]
            sheet.cell(row=row, column=i * 8 + 9).value = data.get("Income_Statement").get(income[i]).get("Quarterly")[
                2]
            sheet.cell(row=row, column=i * 8 + 10).value = data.get("Income_Statement").get(income[i]).get("Quarterly")[
                3]

            # if income[i] != "dates":
            #     # f'[${data.get("currency")} ]#,0'
            #     sheet.cell(row=row, column=i * 8 + 2).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 3).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 4).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 5).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 6).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 7).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 8).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 9).number_format = ccc
        for i in range(len(bal)):
            sheet.cell(row=row, column=i * 8 + 243).value = data.get("Balance_Sheet").get(bal[i]).get("Annual")[0]
            sheet.cell(row=row, column=i * 8 + 244).value = data.get("Balance_Sheet").get(bal[i]).get("Annual")[1]
            sheet.cell(row=row, column=i * 8 + 245).value = data.get("Balance_Sheet").get(bal[i]).get("Annual")[2]
            sheet.cell(row=row, column=i * 8 + 246).value = data.get("Balance_Sheet").get(bal[i]).get("Annual")[3]
            sheet.cell(row=row, column=i * 8 + 247).value = data.get("Balance_Sheet").get(bal[i]).get("Quarterly")[0]
            sheet.cell(row=row, column=i * 8 + 248).value = data.get("Balance_Sheet").get(bal[i]).get("Quarterly")[1]
            sheet.cell(row=row, column=i * 8 + 249).value = data.get("Balance_Sheet").get(bal[i]).get("Quarterly")[2]
            sheet.cell(row=row, column=i * 8 + 250).value = data.get("Balance_Sheet").get(bal[i]).get("Quarterly")[3]
            # if bal[i] != "dates":
            #     sheet.cell(row=row, column=i * 8 + 242).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 243).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 244).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 245).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 246).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 247).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 248).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 249).number_format = ccc
        for i in range(len(cash)):
            sheet.cell(row=row, column=i * 8 + 315).value = data.get("Cash_Flow").get(cash[i]).get("Annual")[0]
            sheet.cell(row=row, column=i * 8 + 316).value = data.get("Cash_Flow").get(cash[i]).get("Annual")[1]
            sheet.cell(row=row, column=i * 8 + 317).value = data.get("Cash_Flow").get(cash[i]).get("Annual")[2]
            sheet.cell(row=row, column=i * 8 + 318).value = data.get("Cash_Flow").get(cash[i]).get("Annual")[3]
            sheet.cell(row=row, column=i * 8 + 319).value = data.get("Cash_Flow").get(cash[i]).get("Quarterly")[0]
            sheet.cell(row=row, column=i * 8 + 320).value = data.get("Cash_Flow").get(cash[i]).get("Quarterly")[1]
            sheet.cell(row=row, column=i * 8 + 321).value = data.get("Cash_Flow").get(cash[i]).get("Quarterly")[2]
            sheet.cell(row=row, column=i * 8 + 322).value = data.get("Cash_Flow").get(cash[i]).get("Quarterly")[3]
            # if cash[i] != "dates":
            #     sheet.cell(row=row, column=i * 8 + 314).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 315).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 316).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 317).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 318).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 319).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 320).number_format = ccc
            #     sheet.cell(row=row, column=i * 8 + 321).number_format = ccc
        book.save(file)
        book.close()
    except Exception as e:
        print(row)
        print(e)
        logging.error(f"to_xlsx_fin() error, row: {row}. {e.__class__.__name__}, {e}")


def get_appl_data_xlsx(file):
    appl_data = []

    book = openpyxl.open(file)
    sheet = book.active

    i = 1
    while sheet[f"A{i}"].value:
        appl_data.append(sheet[f"A{i}"].value)
        i += 1
    book.close()
    logging.info(appl_data)

    return appl_data


def get_yahoo_cookie():
    cookie = None

    response = requests.get(
        "https://fc.yahoo.com", headers=headers, allow_redirects=True
    )

    if not response.cookies:
        return "NO_COOKIES"

    cookie = list(response.cookies)[0]

    return cookie


def get_yahoo_crumb(cookie):
    crumb = None

    crumb_response = requests.get(
        "https://query1.finance.yahoo.com/v1/test/getcrumb",
        headers=headers,
        cookies={cookie.name: cookie.value},
        allow_redirects=True,
    )
    crumb = crumb_response.text

    if crumb is None:
        return "NO_CRUMB"

    return crumb


def main():
    try:
        with open("config.txt") as f:
            config = json.loads(f.read())

        if [i for i in config.keys()] == ['white', 'yellow', 'green', 'description', 'shares', 'timeout',
                                          'cookie'] and all(
            [i in [True, False] or type(i) == int for i in config.values()]):
            print("Config was successfully received")
            logging.info(f"config settings: {config}")
        else:
            print("ERROR! Incorrect config format!")
            logging.error("ERROR! Incorrect config format!")
            print("config example:")
            print("""{
  "white": true,
  "yellow": true,
  "green": true,
  "description": true,
  "shares": true,
  "timeout": 1,
  "cookie": 10
}""")
            return
    except Exception as e:
        print(e)
        print("Config error")
        logging.error("Config error")
        print("config example:")
        print("""{
  "white": true,
  "yellow": true,
  "green": true,
  "description": true,
  "shares": true,
  "timeout": 1,
  "cookie": 10
}""")
        return

    file = input("TYPE XLSX FILENAME AND PRESS ENTER: ")
    # file = "1.xlsx"

    if not file or not os.path.isfile(file):
        print(f"No file: '{file}'")
        logging.error(f"No file: '{file}'")
        return
    appl_data = get_appl_data_xlsx(file)

    f = file.split(".")
    file = f"{f[0]}_result.{f[-1]}"
    file_fin = f"{f[0]}_financials.{f[-1]}"

    xlsx_main(file)
    xlsx_main_fin(file_fin)
    print(appl_data)
    row = 4
    coo = config.get("cookie", 10)
    for stock in appl_data:
        if (row - 4) % coo == 0:
            cookie = get_yahoo_cookie()
            if cookie == "NO_COOKIES":
                time.sleep(5)
                cookie = get_yahoo_cookie()
                if cookie == "NO_COOKIES":
                    print("COOKIE ERROR")
                    return
            crumb = get_yahoo_crumb(cookie)
            if crumb == "NO_CRUMB":
                time.sleep(5)
                crumb = get_yahoo_crumb(cookie)
                if crumb == "NO_CRUMB":
                    print("CRUMB ERROR")
                    return
        inform = Parser(stock, config, cookie, crumb)
        to_xlsx(file, stock, inform.data, row)
        to_xlsx_fin(file_fin, stock, inform.financials, row)

        # with open(f"results_json/result_{stock}.json", "w") as f:
        #     json.dump(inform.data, f)
        # with open(f"results_json/financials_{stock}.json", "w") as f:
        #     json.dump(inform.financials, f)
        row += 1
    print(f"All data is saved in files '{file}' and '{file_fin}'")
    logging.info(f"All data is saved in file '{file}' and '{file_fin}'")


if __name__ == '__main__':
    main()
